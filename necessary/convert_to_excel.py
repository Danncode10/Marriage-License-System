import openpyxl
from openpyxl.drawing.image import Image
from datetime import datetime
import sys
import json
import os

def cm_to_pixels(cm):
    return int((cm / 2.54) * 96)

def to_up(val):
    return str(val).upper().strip() if val else ""

def process_excel(data):
    # Load template from the same directory as the script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, "data/APPLICATION-for-MARRIAGE-LICENSE.xlsx")
    img_path = os.path.join(script_dir, "data/couple_img.png")
    
    wb = openpyxl.load_workbook(template_path)
    
    # Date/Misc
    now = datetime.now()
    day_now = str(now.day)
    month_now = now.strftime("%B").upper()
    year_now = str(now.year)

    # GROOM DATA
    g_first = data.get("gFirst", "")
    g_middle = data.get("gMiddle", "")
    g_last = data.get("gLast", "")
    g_bday = data.get("gBday", "")
    g_age = data.get("gAge", 0)
    g_birth_place = data.get("gBirthPlace", "")
    g_brgy = data.get("gBrgy", "")
    g_town = data.get("gTown", "")
    g_prov = data.get("gProv", "NUEVA VIZCAYA")
    g_country = data.get("gCountry", "PHILIPPINES")
    g_citizen = data.get("gCitizen", "FILIPINO")
    g_status = data.get("gStatus", "SINGLE")
    g_religion = data.get("gReligion", "")
    
    # BRIDE DATA
    b_first = data.get("bFirst", "")
    b_middle = data.get("bMiddle", "")
    b_last = data.get("bLast", "")
    b_bday = data.get("bBday", "")
    b_age = data.get("bAge", 0)
    b_birth_place = data.get("bBirthPlace", "")
    b_brgy = data.get("bBrgy", "")
    b_town = data.get("bTown", "")
    b_prov = data.get("bProv", "NUEVA VIZCAYA")
    b_country = data.get("bCountry", "PHILIPPINES")
    b_citizen = data.get("bCitizen", "FILIPINO")
    b_status = data.get("bStatus", "SINGLE")
    b_religion = data.get("bReligion", "")

    # PARENTS
    g_fath_f = data.get("gFathF", "")
    g_fath_m = data.get("gFathM", "")
    g_fath_l = data.get("gFathL", "")
    g_moth_f = data.get("gMothF", "")
    g_moth_m = data.get("gMothM", "")
    g_moth_l = data.get("gMothL", "")
    
    b_fath_f = data.get("bFathF", "")
    b_fath_m = data.get("bFathM", "")
    b_fath_l = data.get("bFathL", "")
    b_moth_f = data.get("bMothF", "")
    b_moth_m = data.get("bMothM", "")
    b_moth_l = data.get("bMothL", "")

    # GIVERS
    g_giver_f = data.get("gGiverF", "")
    g_giver_m = data.get("gGiverM", "")
    g_giver_l = data.get("gGiverL", "")
    g_giver_relation = data.get("gGiverRelation", "")
    
    b_giver_f = data.get("bGiverF", "")
    b_giver_m = data.get("bGiverM", "")
    b_giver_l = data.get("bGiverL", "")
    b_giver_relation = data.get("bGiverRelation", "")

    # Addresses
    g_full_addr = to_up(f"{g_brgy}, {g_town}, {g_prov}")
    b_full_addr = to_up(f"{b_brgy}, {b_town}, {b_prov}")
    g_town_prov = to_up(f"{g_town}, {g_prov}")
    b_town_prov = to_up(f"{b_town}, {b_prov}")

    is_groom_external = g_town_prov != "SOLANO, NUEVA VIZCAYA"
    is_bride_external = b_town_prov != "SOLANO, NUEVA VIZCAYA"
    needs_back_sheets = is_groom_external or is_bride_external

    # Fill notice sheet image
    if os.path.exists(img_path):
        try:
            couple_img = Image(img_path)
            couple_img.height = cm_to_pixels(3.75)
            couple_img.width = cm_to_pixels(5.73)
            if "Notice" in wb.sheetnames:
                notice_sheet = wb["Notice"]
                notice_sheet.add_image(couple_img, "T11")
        except:
            pass

    app_sheet = wb["APPLICATION"]
    
    # GROOM mapping (APPLICATION)
    app_sheet['B8'] = to_up(g_first)
    app_sheet['B9'] = to_up(g_middle)
    app_sheet['B10'] = to_up(g_last)
    app_sheet['B11'] = to_up(g_bday)
    app_sheet['N11'] = g_age
    app_sheet['B12'] = to_up(g_birth_place) if g_birth_place else g_town_prov
    app_sheet['L12'] = to_up(g_country)
    app_sheet['B13'] = "MALE"
    app_sheet['H13'] = to_up(g_citizen)
    app_sheet['B15'] = g_full_addr
    app_sheet['M15'] = to_up(g_country)
    app_sheet['B16'] = to_up(g_religion)
    app_sheet['B17'] = to_up(g_status)
    app_sheet['B22'] = to_up(g_fath_f)
    app_sheet['H22'] = to_up(g_fath_m)
    app_sheet['L22'] = to_up(g_fath_l)
    app_sheet['B23'] = to_up(g_citizen)
    app_sheet['B25'] = g_full_addr
    app_sheet['M25'] = to_up(g_country)
    app_sheet['B26'] = to_up(g_moth_f)
    app_sheet['G26'] = to_up(g_moth_m)
    app_sheet['K26'] = to_up(g_moth_l)
    app_sheet['B27'] = to_up(g_citizen)
    app_sheet['B29'] = g_full_addr
    app_sheet['M29'] = to_up(g_country)
    app_sheet['B34'] = g_full_addr
    app_sheet['M34'] = to_up(g_country)

    if 18 <= g_age <= 24:
        app_sheet['B30'] = to_up(g_giver_f)
        app_sheet['H30'] = to_up(g_giver_m)
        app_sheet['L30'] = to_up(g_giver_l)
        app_sheet['B31'] = to_up(g_giver_relation)
        app_sheet['B32'] = to_up(g_citizen)

    # FEMALE mapping (APPLICATION)
    app_sheet['U8'] = to_up(b_first)
    app_sheet['U9'] = to_up(b_middle)
    app_sheet['U10'] = to_up(b_last)
    app_sheet['U11'] = to_up(b_bday)
    app_sheet['AF11'] = b_age
    app_sheet['U12'] = to_up(b_birth_place) if b_birth_place else b_town_prov
    app_sheet['AE12'] = to_up(b_country)
    app_sheet['U13'] = "FEMALE"
    app_sheet['Z13'] = to_up(b_citizen)
    app_sheet['U15'] = b_full_addr
    app_sheet['AF15'] = to_up(b_country)
    app_sheet['U16'] = to_up(b_religion)
    app_sheet['U17'] = to_up(b_status)
    app_sheet['U22'] = to_up(b_fath_f)
    app_sheet['Y22'] = to_up(b_fath_m)
    app_sheet['AC22'] = to_up(b_fath_l)
    app_sheet['U23'] = to_up(b_citizen)
    app_sheet['U25'] = b_full_addr
    app_sheet['AF25'] = to_up(b_country)
    app_sheet['U26'] = to_up(b_moth_f)
    app_sheet['Y26'] = to_up(b_moth_m)
    app_sheet['AD26'] = to_up(b_moth_l)
    app_sheet['U27'] = to_up(b_citizen)
    app_sheet['U29'] = b_full_addr
    app_sheet['AF29'] = to_up(b_country)
    app_sheet['U34'] = b_full_addr
    app_sheet['AF34'] = to_up(b_country)

    if 18 <= b_age <= 24:
        app_sheet['U30'] = to_up(b_giver_f)
        app_sheet['Y30'] = to_up(b_giver_m)
        app_sheet['AD30'] = to_up(b_giver_l)
        app_sheet['U31'] = to_up(b_giver_relation)
        app_sheet['U32'] = to_up(b_citizen)

    # Date and Location (APPLICATION)
    app_sheet['B37'] = day_now
    app_sheet['U37'] = day_now
    app_sheet['E37'] = month_now
    app_sheet['W37'] = month_now
    app_sheet['L37'] = year_now
    app_sheet['AD37'] = year_now
    app_sheet['B38'] = "SOLANO, NUEVA VIZCAYA"
    app_sheet['U38'] = "SOLANO, NUEVA VIZCAYA"

    # Logic for sheet visibility based on age (Consent/Advice)
    sheet_name = None
    if (18 <= b_age <= 20 and g_age >= 25):
        sheet_name = "CONSENT F"
    elif (18 <= g_age <= 20 and b_age >= 25):
        sheet_name = "CONSENT M"
    elif (18 <= b_age <= 20 and 18 <= g_age <= 20):
        sheet_name = "CONSENT M&F"
    elif (21 <= b_age <= 24 and g_age >= 25):
        sheet_name = "ADVICE F"
    elif (21 <= g_age <= 24 and b_age >= 25):
        sheet_name = "ADVICE M"
    elif (21 <= b_age <= 24 and 21 <= g_age <= 24):
        sheet_name = "ADVICE M&F"
    elif (21 <= g_age <= 24 and 18 <= b_age <= 20):
        sheet_name = "ADVICE M-CONSENT F"
    elif (21 <= b_age <= 24 and 18 <= g_age <= 20):
        sheet_name = "ADVICE F-CONSENT M"

    # Handle NOTICE tab logic
    if "Notice" in wb.sheetnames:
        notice_sheet = wb["Notice"]
        if is_groom_external and is_bride_external:
            notice_sheet['E44'] = g_full_addr
            notice_sheet['E45'] = b_full_addr
        elif is_groom_external:
            notice_sheet['E44'] = g_full_addr
            notice_sheet['E45'] = ""
        elif is_bride_external:
            notice_sheet['E44'] = b_full_addr
            notice_sheet['E45'] = ""
        else:
            notice_sheet['E44'] = ""
            notice_sheet['E45'] = ""
            notice_sheet['E46'] = ""

    # Filter sheets and visibility
    sheets_to_keep = ["APPLICATION", "Notice"]
    if sheet_name:
        sheets_to_keep.append(sheet_name)
    
    if needs_back_sheets:
        sheets_to_keep.extend(["AddressBACKnotice", "EnvelopeAddress"])
    
    for s in wb.sheetnames:
        if s not in sheets_to_keep:
            del wb[s]
        elif s in ["AddressBACKnotice", "EnvelopeAddress"]:
             wb[s].sheet_state = 'visible'

    # Output to buffer (stdout)
    wb.save(sys.stdout.buffer)

if __name__ == "__main__":
    try:
        input_data = json.load(sys.stdin)
        process_excel(input_data)
    except Exception as e:
        sys.stderr.write(str(e))
        sys.exit(1)
